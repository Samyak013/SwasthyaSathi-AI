import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import random

# Create a new workbook
wb = openpyxl.Workbook()

# Remove default sheet
wb.remove(wb.active)

# Define styles
header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def apply_header_style(cell):
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

def apply_border(cell):
    cell.border = border
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

# ============================================================================
# Setup Instructions
# ============================================================================
ws_instructions = wb.create_sheet("Setup & Login", 0)

setup_content = [
    ["SwasthyaSathi-AI - Complete Login & Testing Guide"],
    [],
    ["🔐 QUICK START - Login Instructions"],
    [],
    ["Step 1: Use any ABHA ID and select contact method (Email or SMS)"],
    ["Step 2: OTP will be sent to the provided email/phone"],
    ["Step 3: In DEVELOPMENT MODE - OTP is logged to browser console"],
    ["Step 4: Enter OTP to access your dashboard"],
    [],
    ["📱 CONTACT METHODS FOR TESTING"],
    ["- Email: Uses real Gmail/SendGrid (configured in production)"],
    ["- SMS: Uses real Twilio/AWS SNS (configured in production)"],
    ["- Dev Mode: OTP printed to console and backend logs"],
    [],
    ["===== DOCTOR LOGIN ====="],
    ["ABHA ID: 22-1234-5678-9012"],
    ["Name: Dr. Rajesh Kumar"],
    ["Email: rajesh.kumar@hospital.com"],
    ["Phone: +91-9876543210"],
    ["OTP Method: Email or SMS"],
    ["Dashboard Access: View Patients, Create Prescriptions, Analytics"],
    [],
    ["===== PATIENT LOGIN 1 ====="],
    ["ABHA ID: 22-1111-2222-3333"],
    ["Name: Priya Sharma"],
    ["Email: priya.sharma@email.com"],
    ["Phone: +91-9123456789"],
    ["OTP Method: Email or SMS"],
    ["Dashboard Access: View Health Records, Prescriptions, Chat with Doctors"],
    [],
    ["===== PATIENT LOGIN 2 ====="],
    ["ABHA ID: 22-4444-5555-6666"],
    ["Name: Arjun Patel"],
    ["Email: arjun.patel@email.com"],
    ["Phone: +91-9876501234"],
    ["OTP Method: Email or SMS"],
    ["Dashboard Access: View Health Records, Prescriptions, Chat with Doctors"],
    [],
    ["===== PHARMACY LOGIN ====="],
    ["ABHA ID: 22-8888-9999-0000"],
    ["Name: Medcare Pharmacy"],
    ["Email: medcare@pharmacy.com"],
    ["Phone: +91-9765432100"],
    ["OTP Method: Email"],
    ["Dashboard Access: Verify Prescriptions, Dispense Medication, Track QR Codes"],
    [],
    ["🔧 DEVELOPMENT ENVIRONMENT"],
    ["When NODE_ENV=development:"],
    ["1. OTP is printed to Server Console (check terminal output)"],
    ["2. OTP is also printed to Browser Console (F12)"],
    ["3. Email/SMS sending is simulated"],
    ["4. All features are available without real API keys"],
    [],
    ["✅ TESTING CHECKLIST"],
    ["[ ] Login as Doctor - Access doctor dashboard"],
    ["[ ] Login as Patient 1 - View health records"],
    ["[ ] Login as Patient 2 - Check prescriptions"],
    ["[ ] Login as Pharmacy - Verify prescriptions"],
    ["[ ] Test Real-time Chat between Doctor and Patient"],
    ["[ ] Create Prescription as Doctor"],
    ["[ ] Receive Prescription as Patient"],
    ["[ ] Dispense Prescription as Pharmacy"],
    ["[ ] Check Analytics Dashboard"],
]

for row_data in setup_content:
    if isinstance(row_data, list) and len(row_data) > 0:
        ws_instructions.append(row_data)
        cell = ws_instructions[ws_instructions.max_row][0]
        
        if '=====' in str(row_data[0]):
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = PatternFill(start_color="764ba2", end_color="764ba2", fill_type="solid")
        elif row_data[0].startswith('Step') or row_data[0].startswith('[ ]'):
            cell.font = Font(bold=True, size=10)
        elif row_data[0].startswith('ABHA') or row_data[0].startswith('Name') or row_data[0].startswith('Email') or row_data[0].startswith('Phone') or row_data[0].startswith('OTP') or row_data[0].startswith('Dashboard'):
            cell.font = Font(size=10, color="333333")
    else:
        ws_instructions.append(row_data)

ws_instructions.column_dimensions['A'].width = 70

# ============================================================================
# 1. USERS SHEET
# ============================================================================
ws_users = wb.create_sheet("Users")
headers = ["ABHA ID", "Name", "Email", "Phone", "Date of Birth", "Gender", "Address", "Role"]
ws_users.append(headers)
for cell in ws_users[1]:
    apply_header_style(cell)

users_data = [
    # Doctors
    ["22-1234-5678-9012", "Dr. Rajesh Kumar", "rajesh.kumar@hospital.com", "+91-9876543210", "1985-05-15", "Male", "Apollo Hospital, Mumbai", "doctor"],
    ["22-2020-3030-4040", "Dr. Neha Singh", "neha.singh@hospital.com", "+91-9876549876", "1987-11-05", "Female", "Fortis Hospital, Hyderabad", "doctor"],
    ["22-3030-4040-5050", "Dr. Anjali Gupta", "anjali.gupta@hospital.com", "+91-9876505050", "1989-09-12", "Female", "Max Healthcare, Ahmedabad", "doctor"],
    
    # Patients
    ["22-1111-2222-3333", "Priya Sharma", "priya.sharma@email.com", "+91-9123456789", "1990-08-22", "Female", "Delhi, India", "patient"],
    ["22-4444-5555-6666", "Arjun Patel", "arjun.patel@email.com", "+91-9876501234", "1988-03-10", "Male", "Bangalore, India", "patient"],
    ["22-5555-6666-7777", "Vikram Desai", "vikram.desai@email.com", "+91-9123498765", "1992-06-18", "Male", "Chennai, India", "patient"],
    
    # Pharmacies
    ["22-8888-9999-0000", "Medcare Pharmacy", "medcare@pharmacy.com", "+91-9765432100", "2015-01-20", "Male", "Pune, Maharashtra", "pharmacy"],
    ["22-7777-8888-9999", "MediPharm Store", "medipharm@pharmacy.com", "+91-9123476543", "2016-05-10", "Female", "Kolkata, West Bengal", "pharmacy"],
]

for row_data in users_data:
    ws_users.append(row_data)
    for cell in ws_users[ws_users.max_row]:
        apply_border(cell)

ws_users.column_dimensions['A'].width = 18
ws_users.column_dimensions['B'].width = 20
ws_users.column_dimensions['C'].width = 28
ws_users.column_dimensions['D'].width = 18
ws_users.column_dimensions['E'].width = 15
ws_users.column_dimensions['F'].width = 10
ws_users.column_dimensions['G'].width = 30
ws_users.column_dimensions['H'].width = 12

# ============================================================================
# 2. DOCTORS SHEET
# ============================================================================
ws_doctors = wb.create_sheet("Doctors")
headers = ["ABHA ID", "Specialization", "Experience (Years)", "Hospital Name", "HPR ID", "Consultation Fee"]
ws_doctors.append(headers)
for cell in ws_doctors[1]:
    apply_header_style(cell)

doctors_data = [
    ["22-1234-5678-9012", "Cardiology", 15, "Apollo Hospital", "HPR001", 500],
    ["22-2020-3030-4040", "Neurology", 12, "Fortis Hospital", "HPR002", 600],
    ["22-3030-4040-5050", "Pediatrics", 10, "Max Healthcare", "HPR003", 400],
]

for row_data in doctors_data:
    ws_doctors.append(row_data)
    for cell in ws_doctors[ws_doctors.max_row]:
        apply_border(cell)

for col, width in [('A', 18), ('B', 15), ('C', 18), ('D', 20), ('E', 12), ('F', 18)]:
    ws_doctors.column_dimensions[col].width = width

# ============================================================================
# 3. PATIENTS SHEET
# ============================================================================
ws_patients = wb.create_sheet("Patients")
headers = ["ABHA ID", "Blood Group", "Height (cm)", "Weight (kg)", "Medical Conditions", "Allergies", "Emergency Contact", "Emergency Phone", "Relation"]
ws_patients.append(headers)
for cell in ws_patients[1]:
    apply_header_style(cell)

patients_data = [
    ["22-1111-2222-3333", "O+", 170, 72, "Hypertension, Diabetes", "Penicillin", "Anita Sharma", "+91-9876543211", "Spouse"],
    ["22-4444-5555-6666", "B+", 175, 75, "Asthma", "Nuts, Shellfish", "Ravi Patel", "+91-9876543212", "Brother"],
    ["22-5555-6666-7777", "AB+", 168, 70, "None", "None", "Sanjana Desai", "+91-9123498766", "Sister"],
]

for row_data in patients_data:
    ws_patients.append(row_data)
    for cell in ws_patients[ws_patients.max_row]:
        apply_border(cell)

for col, width in [('A', 18), ('B', 12), ('C', 14), ('D', 14), ('E', 22), ('F', 18), ('G', 18), ('H', 16), ('I', 12)]:
    ws_patients.column_dimensions[col].width = width

# ============================================================================
# 4. PRESCRIPTIONS SHEET
# ============================================================================
ws_prescriptions = wb.create_sheet("Prescriptions")
headers = ["Prescription ID", "Patient Name", "Patient Phone", "Doctor Name", "Diagnosis", "Symptoms", "Medications (and dosage)", "Lab Tests", "Status"]
ws_prescriptions.append(headers)
for cell in ws_prescriptions[1]:
    apply_header_style(cell)

prescriptions_data = [
    ["RX001", "Priya Sharma", "+91-9123456789", "Dr. Rajesh Kumar", "Hypertension Stage 2", "High BP, Headache", "Amlodipine 5mg (OD), Lisinopril 10mg (OD)", "Blood Pressure Test", "Active"],
    ["RX002", "Arjun Patel", "+91-9876501234", "Dr. Rajesh Kumar", "Asthma Attack", "Shortness of breath", "Salbutamol inhaler (SOS), Fluticasone (BD)", "Lung Function Test", "Active"],
    ["RX003", "Vikram Desai", "+91-9123498765", "Dr. Neha Singh", "Migraine", "Severe headache, Nausea", "Metoprolol 25mg (OD), Sumatriptan (SOS)", "CT Scan Brain", "Dispensed"],
]

for row_data in prescriptions_data:
    ws_prescriptions.append(row_data)
    for cell in ws_prescriptions[ws_prescriptions.max_row]:
        apply_border(cell)

for col, width in [('A', 16), ('B', 16), ('C', 16), ('D', 16), ('E', 18), ('F', 18), ('G', 26), ('H', 18), ('I', 12)]:
    ws_prescriptions.column_dimensions[col].width = width

# ============================================================================
# 5. APPOINTMENTS SHEET
# ============================================================================
ws_appointments = wb.create_sheet("Appointments")
headers = ["Appointment ID", "Patient Name", "Patient Phone", "Doctor Name", "Date", "Time", "Duration (min)", "Type", "Status"]
ws_appointments.append(headers)
for cell in ws_appointments[1]:
    apply_header_style(cell)

tomorrow = (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d")
next_week = (datetime.now() + timedelta(days=7)).strftime("%Y-%m-%d")

appointments_data = [
    ["APT001", "Priya Sharma", "+91-9123456789", "Dr. Rajesh Kumar", tomorrow, "10:00 AM", 30, "Consultation", "Confirmed"],
    ["APT002", "Arjun Patel", "+91-9876501234", "Dr. Rajesh Kumar", next_week, "02:00 PM", 30, "Consultation", "Scheduled"],
    ["APT003", "Vikram Desai", "+91-9123498765", "Dr. Neha Singh", tomorrow, "03:30 PM", 45, "Follow-up", "Confirmed"],
]

for row_data in appointments_data:
    ws_appointments.append(row_data)
    for cell in ws_appointments[ws_appointments.max_row]:
        apply_border(cell)

for col, width in [('A', 16), ('B', 16), ('C', 16), ('D', 16), ('E', 12), ('F', 12), ('G', 16), ('H', 12), ('I', 12)]:
    ws_appointments.column_dimensions[col].width = width

# ============================================================================
# 6. API TEST LINKS SHEET
# ============================================================================
ws_api = wb.create_sheet("API Test Commands")
headers = ["Endpoint", "Method", "URL", "Request Body", "Description"]
ws_api.append(headers)
for cell in ws_api[1]:
    apply_header_style(cell)

api_data = [
    ["Send OTP", "POST", "http://localhost:5000/api/auth/send-otp", '{\"abhaId\": \"22-1234-5678-9012\", \"channel\": \"email\"}', "Send OTP to doctor email"],
    ["Verify OTP", "POST", "http://localhost:5000/api/auth/verify-otp", '{\"abhaId\": \"22-1234-5678-9012\", \"otp\": \"XXXXXX\"}', "Verify OTP and login"],
    ["Get Doctors", "GET", "http://localhost:5000/api/doctors", "N/A", "List all doctors"],
    ["Get User", "GET", "http://localhost:5000/api/users/{userId}", "N/A", "Get user details"],
    ["Create Prescription", "POST", "http://localhost:5000/api/prescriptions", "{...prescription data}", "Create new prescription"],
]

for row_data in api_data:
    ws_api.append(row_data)
    for cell in ws_api[ws_api.max_row]:
        apply_border(cell)

for col, width in [('A', 18), ('B', 10), ('C', 40), ('D', 40), ('E', 30)]:
    ws_api.column_dimensions[col].width = width

# Save the workbook
file_path = "c:\\Users\\samya\\OneDrive\\Desktop\\Swashtya-sathi-ai\\SwasthyaSathi_Login_Guide.xlsx"
wb.save(file_path)
print(f"✅ Login Guide created: {file_path}")
print("\nTest Credentials Summary:")
print("\n👨‍⚕️ DOCTORS (1 available)")
print("  - Dr. Rajesh Kumar: 22-1234-5678-9012 (rajesh.kumar@hospital.com)")
print("\n👨‍⚕️ PATIENTS (3 available)")
print("  - Priya Sharma: 22-1111-2222-3333 (priya.sharma@email.com)")
print("  - Arjun Patel: 22-4444-5555-6666 (arjun.patel@email.com)")
print("  - Vikram Desai: 22-5555-6666-7777 (vikram.desai@email.com)")
print("\n💊 PHARMACIES (2 available)")
print("  - Medcare Pharmacy: 22-8888-9999-0000 (medcare@pharmacy.com)")
print("  - MediPharm Store: 22-7777-8888-9999 (medipharm@pharmacy.com)")
