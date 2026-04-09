import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import random

# Create a new workbook
wb = openpyxl.Workbook()

# Remove default sheet
wb.remove(wb.active)

# Define styles
header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
subheader_fill = PatternFill(start_color="E8E8F0", end_color="E8E8F0", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def apply_header_style(cell):
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

def apply_border(cell):
    cell.border = border
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

# ============================================================================
# 1. USERS SHEET
# ============================================================================
ws_users = wb.create_sheet("Users")
headers = ["ABHA ID", "Name", "Email", "Phone", "Date of Birth", "Gender", "Address", "Role"]
ws_users.append(headers)
for cell in ws_users[1]:
    apply_header_style(cell)

users_data = [
    ["22-1234-5678-9012", "Dr. Rajesh Kumar", "rajesh.kumar@hospital.com", "9876543210", "1985-05-15", "Male", "Mumbai, India", "doctor"],
    ["22-1111-2222-3333", "Priya Sharma", "priya.sharma@email.com", "9123456789", "1990-08-22", "Female", "Delhi, India", "patient"],
    ["22-4444-5555-6666", "Arjun Patel", "arjun.patel@email.com", "9876501234", "1988-03-10", "Male", "Bangalore, India", "patient"],
    ["22-8888-9999-0000", "Medcare Pharmacy", "medcare@pharmacy.com", "9765432100", "2015-01-20", "Male", "Pune, India", "pharmacy"],
    ["22-2020-3030-4040", "Dr. Neha Singh", "neha.singh@hospital.com", "9876549876", "1987-11-05", "Female", "Hyderabad, India", "doctor"],
    ["22-5555-6666-7777", "Vikram Desai", "vikram.desai@email.com", "9123498765", "1992-06-18", "Male", "Chennai, India", "patient"],
    ["22-7777-8888-9999", "MediPharm Store", "medipharm@pharmacy.com", "9123476543", "2016-05-10", "Female", "Kolkata, India", "pharmacy"],
    ["22-3030-4040-5050", "Dr. Anjali Gupta", "anjali.gupta@hospital.com", "9876505050", "1989-09-12", "Female", "Ahmedabad, India", "doctor"],
]

for row_data in users_data:
    ws_users.append(row_data)
    for cell in ws_users[ws_users.max_row]:
        apply_border(cell)

ws_users.column_dimensions['A'].width = 18
ws_users.column_dimensions['B'].width = 20
ws_users.column_dimensions['C'].width = 28
ws_users.column_dimensions['D'].width = 15
ws_users.column_dimensions['E'].width = 15
ws_users.column_dimensions['F'].width = 10
ws_users.column_dimensions['G'].width = 25
ws_users.column_dimensions['H'].width = 12

# ============================================================================
# 2. DOCTORS SHEET
# ============================================================================
ws_doctors = wb.create_sheet("Doctors")
headers = ["ABHA ID", "Specialization", "Experience (Years)", "Hospital Name", "HPR ID", "Consultation Fee"]
ws_doctors.append(headers)
for cell in ws_doctors[1]:
    apply_header_style(cell)

doctors_data = [
    ["22-1234-5678-9012", "Cardiology", 15, "Apollo Hospital", "HPR001", 500],
    ["22-2020-3030-4040", "Neurology", 12, "Fortis Hospital", "HPR002", 600],
    ["22-3030-4040-5050", "Pediatrics", 10, "Max Healthcare", "HPR003", 400],
]

for row_data in doctors_data:
    ws_doctors.append(row_data)
    for cell in ws_doctors[ws_doctors.max_row]:
        apply_border(cell)

ws_doctors.column_dimensions['A'].width = 18
ws_doctors.column_dimensions['B'].width = 15
ws_doctors.column_dimensions['C'].width = 18
ws_doctors.column_dimensions['D'].width = 20
ws_doctors.column_dimensions['E'].width = 12
ws_doctors.column_dimensions['F'].width = 18

# ============================================================================
# 3. PATIENTS SHEET
# ============================================================================
ws_patients = wb.create_sheet("Patients")
headers = ["ABHA ID", "Blood Group", "Height (cm)", "Weight (kg)", "Medical Conditions", "Allergies", "Emergency Contact Name", "Emergency Contact Phone", "Emergency Contact Relation"]
ws_patients.append(headers)
for cell in ws_patients[1]:
    apply_header_style(cell)

patients_data = [
    ["22-1111-2222-3333", "O+", 170, 72, "Hypertension, Diabetes", "Penicillin, Aspirin", "Anita Sharma", "9876543211", "Spouse"],
    ["22-4444-5555-6666", "B+", 175, 75, "Asthma", "Nuts", "Ravi Patel", "9876543212", "Brother"],
    ["22-5555-6666-7777", "AB+", 168, 70, "None", "None", "Sanjana Desai", "9123498766", "Sister"],
]

for row_data in patients_data:
    ws_patients.append(row_data)
    for cell in ws_patients[ws_patients.max_row]:
        apply_border(cell)

ws_patients.column_dimensions['A'].width = 18
for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
    ws_patients.column_dimensions[col].width = 20

# ============================================================================
# 4. PHARMACIES SHEET
# ============================================================================
ws_pharmacies = wb.create_sheet("Pharmacies")
headers = ["ABHA ID", "License Number", "Location", "Opening Time", "Closing Time", "Verified"]
ws_pharmacies.append(headers)
for cell in ws_pharmacies[1]:
    apply_header_style(cell)

pharmacies_data = [
    ["22-8888-9999-0000", "LIC001", "Pune, Maharashtra", "09:00", "21:00", "Yes"],
    ["22-7777-8888-9999", "LIC002", "Kolkata, West Bengal", "08:00", "22:00", "Yes"],
]

for row_data in pharmacies_data:
    ws_pharmacies.append(row_data)
    for cell in ws_pharmacies[ws_pharmacies.max_row]:
        apply_border(cell)

ws_pharmacies.column_dimensions['A'].width = 18
ws_pharmacies.column_dimensions['B'].width = 16
ws_pharmacies.column_dimensions['C'].width = 25
ws_pharmacies.column_dimensions['D'].width = 14
ws_pharmacies.column_dimensions['E'].width = 14
ws_pharmacies.column_dimensions['F'].width = 10

# ============================================================================
# 5. PRESCRIPTIONS SHEET
# ============================================================================
ws_prescriptions = wb.create_sheet("Prescriptions")
headers = ["Prescription ID", "Patient ABHA ID", "Doctor ABHA ID", "Diagnosis", "Symptoms", "Medications", "Lab Tests", "Valid Until", "Status", "Notes"]
ws_prescriptions.append(headers)
for cell in ws_prescriptions[1]:
    apply_header_style(cell)

prescriptions_data = [
    ["RX001", "22-1111-2222-3333", "22-1234-5678-9012", "Hypertension Stage 2", "High BP, Headache", "Amlodipine 5mg, Lisinopril 10mg", "Blood Pressure Test", "2024-05-10", "Active", "Regular monitoring"],
    ["RX002", "22-4444-5555-6666", "22-1234-5678-9012", "Asthma Attack", "Shortness of breath, Wheezing", "Salbutamol inhaler, Fluticasone", "Lung Function Test", "2024-04-20", "Active", "Use inhaler as needed"],
    ["RX003", "22-5555-6666-7777", "22-2020-3030-4040", "Migraine", "Severe headache, Nausea", "Metoprolol 25mg, Sumatriptan", "CT Scan Brain", "2024-05-15", "Active", "Avoid triggers"],
]

for row_data in prescriptions_data:
    ws_prescriptions.append(row_data)
    for cell in ws_prescriptions[ws_prescriptions.max_row]:
        apply_border(cell)

for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
    ws_prescriptions.column_dimensions[col].width = 22

# ============================================================================
# 6. APPOINTMENTS SHEET
# ============================================================================
ws_appointments = wb.create_sheet("Appointments")
headers = ["Appointment ID", "Patient ABHA ID", "Doctor ABHA ID", "Scheduled Date", "Time", "Duration (min)", "Type", "Status", "Notes"]
ws_appointments.append(headers)
for cell in ws_appointments[1]:
    apply_header_style(cell)

tomorrow = (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d")
next_week = (datetime.now() + timedelta(days=7)).strftime("%Y-%m-%d")

appointments_data = [
    ["APT001", "22-1111-2222-3333", "22-1234-5678-9012", tomorrow, "10:00 AM", 30, "Consultation", "Confirmed", "Follow-up for BP"],
    ["APT002", "22-4444-5555-6666", "22-1234-5678-9012", next_week, "02:00 PM", 30, "Consultation", "Scheduled", "Asthma review"],
    ["APT003", "22-5555-6666-7777", "22-2020-3030-4040", tomorrow, "03:30 PM", 45, "Consultation", "Confirmed", "Migraine management"],
]

for row_data in appointments_data:
    ws_appointments.append(row_data)
    for cell in ws_appointments[ws_appointments.max_row]:
        apply_border(cell)

for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
    ws_appointments.column_dimensions[col].width = 18

# ============================================================================
# 7. HEALTH RECORDS SHEET
# ============================================================================
ws_health_records = wb.create_sheet("Health Records")
headers = ["Record ID", "Patient ABHA ID", "Doctor ABHA ID", "Record Type", "Title", "Description", "Date Created"]
ws_health_records.append(headers)
for cell in ws_health_records[1]:
    apply_header_style(cell)

today = datetime.now().strftime("%Y-%m-%d")

records_data = [
    ["HR001", "22-1111-2222-3333", "22-1234-5678-9012", "Lab Report", "Blood Test Report", "Complete blood count with normal values", today],
    ["HR002", "22-4444-5555-6666", "22-1234-5678-9012", "Lab Report", "Chest X-Ray Report", "Normal chest with no abnormalities", today],
    ["HR003", "22-5555-6666-7777", "22-2020-3030-4040", "Imaging", "CT Scan Brain Report", "No lesions or abnormalities detected", today],
    ["HR004", "22-1111-2222-3333", "22-1234-5678-9012", "Vital Signs", "BP & HR Monitoring", "BP: 140/90, HR: 78", today],
]

for row_data in records_data:
    ws_health_records.append(row_data)
    for cell in ws_health_records[ws_health_records.max_row]:
        apply_border(cell)

for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
    ws_health_records.column_dimensions[col].width = 18

# ============================================================================
# 8. LOGIN CREDENTIALS SHEET
# ============================================================================
ws_credentials = wb.create_sheet("Login Credentials")
headers = ["Role", "ABHA ID", "Name", "Email", "Phone", "OTP Method", "Notes"]
ws_credentials.append(headers)
for cell in ws_credentials[1]:
    apply_header_style(cell)

credentials_data = [
    ["Doctor", "22-1234-5678-9012", "Dr. Rajesh Kumar", "rajesh.kumar@hospital.com", "9876543210", "Email/SMS", "Senior Cardiologist"],
    ["Doctor", "22-2020-3030-4040", "Dr. Neha Singh", "neha.singh@hospital.com", "9876549876", "Email/SMS", "Neurology Specialist"],
    ["Doctor", "22-3030-4040-5050", "Dr. Anjali Gupta", "anjali.gupta@hospital.com", "9876505050", "Email/SMS", "Pediatrician"],
    ["Patient", "22-1111-2222-3333", "Priya Sharma", "priya.sharma@email.com", "9123456789", "Email/SMS", "Regular Patient"],
    ["Patient", "22-4444-5555-6666", "Arjun Patel", "arjun.patel@email.com", "9876501234", "Email/SMS", "Regular Patient"],
    ["Patient", "22-5555-6666-7777", "Vikram Desai", "vikram.desai@email.com", "9123498765", "Email/SMS", "Regular Patient"],
    ["Pharmacy", "22-8888-9999-0000", "Medcare Pharmacy", "medcare@pharmacy.com", "9765432100", "Email", "Verified Pharmacy"],
    ["Pharmacy", "22-7777-8888-9999", "MediPharm Store", "medipharm@pharmacy.com", "9123476543", "Email", "Verified Pharmacy"],
]

for row_data in credentials_data:
    ws_credentials.append(row_data)
    for cell in ws_credentials[ws_credentials.max_row]:
        apply_border(cell)

for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
    ws_credentials.column_dimensions[col].width = 20

# ============================================================================
# 9. SETUP INSTRUCTIONS SHEET
# ============================================================================
ws_instructions = wb.create_sheet("Setup Instructions", 0)

instructions = [
    ["SwasthyaSathi-AI - Sample Data Setup Guide"],
    [],
    ["QUICK START"],
    ["1. Use any ABHA ID from the Users sheet to login"],
    ["2. OTP will be sent via Email or SMS (logged to console in dev mode)"],
    ["3. Use the OTP displayed in console to complete login"],
    [],
    ["TEST USERS"],
    ["Doctor: 22-1234-5678-9012 (Dr. Rajesh Kumar)"],
    ["Patient: 22-1111-2222-3333 (Priya Sharma)"],
    ["Pharmacy: 22-8888-9999-0000 (Medcare Pharmacy)"],
    [],
    ["DASHBOARD ACCESS"],
    ["- Doctors: View patients, create prescriptions, analytics"],
    ["- Patients: View health records, receive prescriptions, chat with doctors"],
    ["- Pharmacy: Dispense prescriptions, verify QR codes"],
    [],
    ["FEATURES TO TEST"],
    ["1. User Registration and Login with OTP"],
    ["2. Doctor Dashboard - Patient list and prescription creation"],
    ["3. Patient Dashboard - Health records and prescriptions"],
    ["4. Pharmacy Portal - Prescription verification and dispensing"],
    ["5. Real-time Chat between users (use console logs)"],
    ["6. Appointment Booking and Management"],
    ["7. Health Reminders and Notifications"],
    ["8. AI Features (Diagnosis suggestions, Drug interactions)"],
    [],
    ["OTP SETTINGS"],
    ["- OTP Expiry: 5 minutes"],
    ["- Supports: Email, SMS, or Both"],
    ["- Max Attempts: 3"],
    [],
    ["DATABASE SEEDING"],
    ["All sample data is automatically seeded on first app load"],
    ["To refresh: Restart the application"],
]

for row_data in instructions:
    if isinstance(row_data, list) and len(row_data) > 0:
        ws_instructions.append(row_data)
        if row_data[0] and not row_data[0].startswith(' '):
            for cell in ws_instructions[ws_instructions.max_row]:
                if ws_instructions[ws_instructions.max_row].index(cell) == 0:
                    if any(x in row_data[0] for x in ["Guide", "TEST", "Access", "Settings", "Seeding"]):
                        cell.font = Font(bold=True, size=12, color="667eea")
                    elif row_data[0].startswith("1") or row_data[0].startswith("2") or row_data[0].startswith("3"):
                        cell.font = Font(bold=True, size=11)
    else:
        ws_instructions.append(row_data)

ws_instructions.column_dimensions['A'].width = 60

# Save the workbook
file_path = "c:\\Users\\samya\\OneDrive\\Desktop\\Swashtya-sathi-ai\\SwasthyaSathi_Sample_Data.xlsx"
wb.save(file_path)
print(f"✅ Excel file created successfully: {file_path}")
print("\nSample Data Summary:")
print(f"  - Users: {len(users_data)}")
print(f"  - Doctors: {len(doctors_data)}")
print(f"  - Patients: {len(patients_data)}")
print(f"  - Pharmacies: {len(pharmacies_data)}")
print(f"  - Prescriptions: {len(prescriptions_data)}")
print(f"  - Appointments: {len(appointments_data)}")
print(f"  - Health Records: {len(records_data)}")
